"""
Data Mining Agent - Agente de Extracción de Datos Avanzada
Proporciona capacidades sofisticadas de extracción, transformación y análisis de datos
desde múltiples fuentes web con procesamiento inteligente.

Características principales:
- Extracción de datos estructurados desde APIs y sitios web
- Transformación y limpieza automática de datos
- Análisis estadístico y detección de patrones
- Validación y verificación de integridad de datos
- Exportación en múltiples formatos (JSON, CSV, XML)
- Programación de extracciones periódicas

Autor: Data Mining Agent
Versión: 1.0.0
"""

import asyncio
import json
import csv
import xml.etree.ElementTree as ET
import re
import time
import hashlib
import logging
import sqlite3
import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Union, Set, Tuple, Callable, Iterator
from urllib.parse import urlparse, urljoin
from dataclasses import dataclass, field, asdict
from enum import Enum
import requests
from bs4 import BeautifulSoup
import sqlite3
import threading
from concurrent.futures import ThreadPoolExecutor

try:
    from ..base_agent_wrapper import BaseAgentWrapper
    from ..search_engine_agent import SearchEngineAgent, SearchSource
except ImportError:
    BaseAgentWrapper = object
    SearchEngineAgent = object
    SearchSource = Enum


class DataSourceType(Enum):
    """Tipos de fuentes de datos"""
    WEB_API = "web_api"          # APIs web
    WEB_SCRAPING = "web_scraping" # Scraping de sitios web
    DATABASE = "database"        # Bases de datos
    RSS_FEED = "rss_feed"        # Feeds RSS
    SOCIAL_MEDIA = "social_media" # Redes sociales
    FILE_DOWNLOAD = "file_download" # Descarga de archivos
    STREAMING = "streaming"      # Datos en tiempo real


class DataFormat(Enum):
    """Formatos de datos soportados"""
    JSON = "json"
    CSV = "csv"
    XML = "xml"
    EXCEL = "excel"
    DATABASE = "database"
    PARQUET = "parquet"


class DataQuality(Enum):
    """Niveles de calidad de datos"""
    EXCELLENT = "excellent"    # Datos completos, consistentes y verificados
    GOOD = "good"             # Datos de buena calidad con mínimos problemas
    FAIR = "fair"             # Datos aceptables con algunos problemas menores
    POOR = "poor"             # Datos problemáticos que requieren limpieza
    INVALID = "invalid"       # Datos inválidos o corruptos


@dataclass
class DataRecord:
    """Registro de datos individual"""
    id: str
    source_url: str
    data: Dict[str, Any]
    extracted_at: float
    quality_score: float
    validation_errors: List[str]
    metadata: Dict[str, Any] = field(default_factory=dict)


@dataclass
class DataSet:
    """Conjunto de datos extraídos"""
    name: str
    description: str
    source_type: DataSourceType
    records: List[DataRecord]
    total_records: int
    quality_assessment: DataQuality
    schema: Dict[str, str]  # field_name -> data_type
    extraction_config: Dict[str, Any]
    created_at: float
    last_updated: float
    metadata: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ExtractionJob:
    """Trabajo de extracción de datos"""
    job_id: str
    source_config: Dict[str, Any]
    schedule_config: Dict[str, Any]
    status: str  # "pending", "running", "completed", "failed"
    progress: float  # 0.0 - 1.0
    created_at: float
    started_at: Optional[float]
    completed_at: Optional[float]
    error_message: Optional[str]
    result_dataset: Optional[DataSet]


class DataMiningAgent(BaseAgentWrapper if BaseAgentWrapper else object):
    """
    Agente especializado en extracción y análisis de datos
    Proporciona capacidades avanzadas de mining, transformación y análisis
    """
    
    def __init__(self):
        super().__init__() if BaseAgentWrapper else None
        
        self.name = "data_mining_agent"
        self.description = "Agente de extracción de datos avanzada con análisis y validación"
        self.version = "1.0.0"
        self.logger = logging.getLogger(__name__)
        
        # Inicializar motor de búsqueda
        self.search_engine = SearchEngineAgent()
        
        # Base de datos interna para almacenamiento
        self._init_database()
        
        # Configuración de extracción
        self.config = {
            "max_concurrent_extractions": 5,
            "timeout_seconds": 30,
            "retry_attempts": 3,
            "batch_size": 100,
            "quality_threshold": 0.7,
            "enable_data_validation": True,
            "enable_deduplication": True,
            "cache_extractions": True,
            "supported_formats": ["json", "csv", "xml", "excel"]
        }
        
        # Cache para optimizaciones
        self._extraction_cache = {}
        self._schema_cache = {}
        
        # Thread pool para extracciones concurrentes
        self._thread_pool = ThreadPoolExecutor(max_workers=self.config["max_concurrent_extractions"])
        
        # Trabajos activos
        self._active_jobs = {}
        self._job_lock = threading.Lock()
    
    def extract_data(
        self,
        source_config: Dict[str, Any],
        output_format: DataFormat = DataFormat.JSON,
        enable_validation: bool = True,
        **kwargs
    ) -> DataSet:
        """
        Extrae datos desde una fuente específica
        
        Args:
            source_config: Configuración de la fuente de datos
            output_format: Formato de salida deseado
            enable_validation: Si habilitar validación de datos
            **kwargs: Parámetros adicionales
            
        Returns:
            DataSet con los datos extraídos
        """
        try:
            self.logger.info(f"Iniciando extracción de datos desde {source_config.get('type', 'unknown')}")
            
            # Validar configuración de fuente
            if not self._validate_source_config(source_config):
                raise ValueError("Configuración de fuente inválida")
            
            # Determinar tipo de fuente
            source_type = DataSourceType(source_config.get("type", "web_scraping"))
            
            # Realizar extracción según el tipo
            if source_type == DataSourceType.WEB_API:
                records = self._extract_from_api(source_config, **kwargs)
            elif source_type == DataSourceType.WEB_SCRAPING:
                records = self._extract_from_web(source_config, **kwargs)
            elif source_type == DataSourceType.RSS_FEED:
                records = self._extract_from_rss(source_config, **kwargs)
            elif source_type == DataSourceType.FILE_DOWNLOAD:
                records = self._extract_from_file(source_config, **kwargs)
            else:
                raise ValueError(f"Tipo de fuente no soportado: {source_type}")
            
            if not records:
                raise ValueError("No se extrajeron datos de la fuente")
            
            # Validar y limpiar datos
            if enable_validation:
                records = self._validate_and_clean_records(records)
            
            # Evaluar calidad
            quality_assessment = self._assess_data_quality(records)
            
            # Generar esquema
            schema = self._infer_schema(records)
            
            # Crear dataset
            dataset = DataSet(
                name=source_config.get("name", "Extracción sin nombre"),
                description=source_config.get("description", ""),
                source_type=source_type,
                records=records,
                total_records=len(records),
                quality_assessment=quality_assessment,
                schema=schema,
                extraction_config=source_config,
                created_at=time.time(),
                last_updated=time.time(),
                metadata={
                    "extraction_method": source_type.value,
                    "validation_enabled": enable_validation,
                    "output_format": output_format.value
                }
            )
            
            # Guardar en cache si está habilitado
            if self.config["cache_extractions"]:
                cache_key = self._generate_cache_key(source_config)
                self._extraction_cache[cache_key] = dataset
            
            self.logger.info(f"Extracción completada: {len(records)} registros")
            return dataset
            
        except Exception as e:
            self.logger.error(f"Error en extracción de datos: {e}")
            raise
    
    def extract_batch(
        self,
        source_configs: List[Dict[str, Any]],
        output_format: DataFormat = DataFormat.JSON,
        max_concurrent: int = 3
    ) -> List[DataSet]:
        """
        Extrae datos de múltiples fuentes de forma concurrente
        
        Args:
            source_configs: Lista de configuraciones de fuentes
            output_format: Formato de salida
            max_concurrent: Máximo número de extracciones concurrentes
            
        Returns:
            Lista de DataSets extraídos
        """
        self.logger.info(f"Iniciando extracción batch de {len(source_configs)} fuentes")
        
        # Limitar concurrencia
        max_concurrent = min(max_concurrent, self.config["max_concurrent_extractions"])
        
        # Ejecutar extracciones concurrentemente
        datasets = []
        with ThreadPoolExecutor(max_workers=max_concurrent) as executor:
            future_to_config = {
                executor.submit(
                    self.extract_data, config, output_format
                ): config for config in source_configs
            }
            
            for future in future_to_config:
                try:
                    dataset = future.result()
                    datasets.append(dataset)
                except Exception as e:
                    config = future_to_config[future]
                    self.logger.error(f"Error en extracción batch para {config.get('url', 'unknown')}: {e}")
        
        self.logger.info(f"Extracción batch completada: {len(datasets)} datasets exitosos")
        return datasets
    
    def analyze_dataset(self, dataset: DataSet) -> Dict[str, Any]:
        """
        Analiza un dataset completo
        
        Args:
            dataset: Dataset a analizar
            
        Returns:
            Dict con análisis detallado del dataset
        """
        try:
            self.logger.info(f"Analizando dataset: {dataset.name}")
            
            analysis = {
                "dataset_info": {
                    "name": dataset.name,
                    "description": dataset.description,
                    "source_type": dataset.source_type.value,
                    "total_records": dataset.total_records,
                    "quality_assessment": dataset.quality_assessment.value,
                    "created_at": dataset.created_at,
                    "last_updated": dataset.last_updated
                },
                "schema_analysis": self._analyze_schema(dataset.schema),
                "data_quality_analysis": self._detailed_quality_analysis(dataset.records),
                "statistical_analysis": self._statistical_analysis(dataset.records),
                "pattern_detection": self._detect_patterns(dataset.records),
                "data_completeness": self._analyze_completeness(dataset.records),
                "recommendations": self._generate_analysis_recommendations(dataset)
            }
            
            return analysis
            
        except Exception as e:
            self.logger.error(f"Error analizando dataset: {e}")
            return {"error": str(e)}
    
    def transform_dataset(
        self,
        dataset: DataSet,
        transformations: List[Dict[str, Any]],
        output_format: DataFormat = DataFormat.JSON
    ) -> DataSet:
        """
        Transforma un dataset aplicando reglas específicas
        
        Args:
            dataset: Dataset a transformar
            transformations: Lista de transformaciones a aplicar
            output_format: Formato de salida
            
        Returns:
            DataSet transformado
        """
        try:
            self.logger.info(f"Transformando dataset: {dataset.name}")
            
            transformed_records = []
            
            for record in dataset.records:
                transformed_record = record
                
                # Aplicar cada transformación
                for transform in transformations:
                    transformed_record = self._apply_transformation(
                        transformed_record, transform
                    )
                
                transformed_records.append(transformed_record)
            
            # Crear nuevo dataset transformado
            transformed_dataset = DataSet(
                name=f"{dataset.name} - Transformado",
                description=f"Versión transformada de {dataset.description}",
                source_type=dataset.source_type,
                records=transformed_records,
                total_records=len(transformed_records),
                quality_assessment=dataset.quality_assessment,
                schema=self._infer_schema(transformed_records),
                extraction_config=dataset.extraction_config,
                created_at=dataset.created_at,
                last_updated=time.time(),
                metadata={
                    **dataset.metadata,
                    "transformations_applied": [t.get("type", "unknown") for t in transformations],
                    "original_dataset": dataset.name
                }
            )
            
            self.logger.info(f"Transformación completada: {len(transformed_records)} registros")
            return transformed_dataset
            
        except Exception as e:
            self.logger.error(f"Error transformando dataset: {e}")
            raise
    
    def export_dataset(
        self,
        dataset: DataSet,
        output_path: str,
        output_format: DataFormat = DataFormat.JSON,
        **kwargs
    ) -> str:
        """
        Exporta un dataset a archivo
        
        Args:
            dataset: Dataset a exportar
            output_path: Ruta del archivo de salida
            output_format: Formato de exportación
            **kwargs: Parámetros adicionales de exportación
            
        Returns:
            Ruta del archivo exportado
        """
        try:
            self.logger.info(f"Exportando dataset {dataset.name} a {output_format.value}")
            
            if output_format == DataFormat.JSON:
                return self._export_to_json(dataset, output_path, **kwargs)
            elif output_format == DataFormat.CSV:
                return self._export_to_csv(dataset, output_path, **kwargs)
            elif output_format == DataFormat.XML:
                return self._export_to_xml(dataset, output_path, **kwargs)
            elif output_format == DataFormat.EXCEL:
                return self._export_to_excel(dataset, output_path, **kwargs)
            elif output_format == DataFormat.DATABASE:
                return self._export_to_database(dataset, output_path, **kwargs)
            else:
                raise ValueError(f"Formato de exportación no soportado: {output_format}")
                
        except Exception as e:
            self.logger.error(f"Error exportando dataset: {e}")
            raise
    
    def schedule_extraction(
        self,
        source_config: Dict[str, Any],
        schedule_config: Dict[str, Any]
    ) -> str:
        """
        Programa una extracción periódica
        
        Args:
            source_config: Configuración de la fuente
            schedule_config: Configuración del cronograma
            
        Returns:
            ID del trabajo programado
        """
        job_id = f"job_{int(time.time())}_{hashlib.md5(str(source_config).encode()).hexdigest()[:8]}"
        
        job = ExtractionJob(
            job_id=job_id,
            source_config=source_config,
            schedule_config=schedule_config,
            status="pending",
            progress=0.0,
            created_at=time.time(),
            started_at=None,
            completed_at=None,
            error_message=None,
            result_dataset=None
        )
        
        with self._job_lock:
            self._active_jobs[job_id] = job
        
        self.logger.info(f"Extracción programada: {job_id}")
        return job_id
    
    def _extract_from_api(self, source_config: Dict[str, Any], **kwargs) -> List[DataRecord]:
        """Extrae datos desde una API web"""
        
        records = []
        
        try:
            # Extraer configuración de API
            base_url = source_config.get("url")
            headers = source_config.get("headers", {})
            params = source_config.get("params", {})
            auth = source_config.get("auth", {})
            
            # Realizar request
            response = requests.get(
                base_url,
                headers=headers,
                params=params,
                auth=auth,
                timeout=self.config["timeout_seconds"]
            )
            response.raise_for_status()
            
            # Parsear respuesta
            data = response.json() if response.headers.get('content-type', '').startswith('application/json') else response.text
            
            # Convertir a registros
            if isinstance(data, dict) and 'results' in data:
                items = data['results']
            elif isinstance(data, list):
                items = data
            elif isinstance(data, dict):
                items = [data]
            else:
                items = []
            
            for i, item in enumerate(items):
                record = DataRecord(
                    id=f"api_record_{int(time.time())}_{i}",
                    source_url=base_url,
                    data=item if isinstance(item, dict) else {"content": item},
                    extracted_at=time.time(),
                    quality_score=0.8,  # APIs generalmente son fuentes confiables
                    validation_errors=[]
                )
                records.append(record)
            
        except Exception as e:
            self.logger.error(f"Error extrayendo desde API: {e}")
            raise
        
        return records
    
    def _extract_from_web(self, source_config: Dict[str, Any], **kwargs) -> List[DataRecord]:
        """Extrae datos desde sitios web mediante scraping"""
        
        records = []
        
        try:
            url = source_config.get("url")
            selectors = source_config.get("selectors", {})
            
            # Realizar request
            response = requests.get(url, timeout=self.config["timeout_seconds"])
            response.raise_for_status()
            
            # Parsear HTML
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Extraer datos según selectores
            for selector_name, selector_config in selectors.items():
                elements = soup.select(selector_config.get("css_selector", ""))
                
                for i, element in enumerate(elements):
                    data = {selector_name: element.get_text(strip=True)}
                    
                    # Extraer atributos adicionales si se especifica
                    if "attributes" in selector_config:
                        for attr_name, attr_selector in selector_config["attributes"].items():
                            attr_element = element.select_one(attr_selector)
                            if attr_element:
                                data[f"{selector_name}_{attr_name}"] = attr_element.get(attr_name, "")
                    
                    record = DataRecord(
                        id=f"web_record_{int(time.time())}_{i}_{selector_name}",
                        source_url=url,
                        data=data,
                        extracted_at=time.time(),
                        quality_score=0.6,  # Scraping puede ser menos confiable
                        validation_errors=[]
                    )
                    records.append(record)
            
        except Exception as e:
            self.logger.error(f"Error extrayendo desde web: {e}")
            raise
        
        return records
    
    def _extract_from_rss(self, source_config: Dict[str, Any], **kwargs) -> List[DataRecord]:
        """Extrae datos desde feeds RSS"""
        
        records = []
        
        try:
            import feedparser
            
            url = source_config.get("url")
            feed = feedparser.parse(url)
            
            for i, entry in enumerate(feed.entries):
                data = {
                    "title": entry.get("title", ""),
                    "link": entry.get("link", ""),
                    "description": entry.get("description", ""),
                    "published": entry.get("published", ""),
                    "author": entry.get("author", "")
                }
                
                record = DataRecord(
                    id=f"rss_record_{int(time.time())}_{i}",
                    source_url=url,
                    data=data,
                    extracted_at=time.time(),
                    quality_score=0.7,
                    validation_errors=[]
                )
                records.append(record)
                
        except ImportError:
            self.logger.warning("feedparser no disponible, usando implementación simplificada")
            return self._extract_rss_simplified(source_config)
        except Exception as e:
            self.logger.error(f"Error extrayendo desde RSS: {e}")
            raise
        
        return records
    
    def _extract_rss_simplified(self, source_config: Dict[str, Any]) -> List[DataRecord]:
        """Implementación simplificada de extracción RSS"""
        
        records = []
        
        try:
            url = source_config.get("url")
            response = requests.get(url, timeout=self.config["timeout_seconds"])
            response.raise_for_status()
            
            # Parse XML manualmente (simplificado)
            soup = BeautifulSoup(response.content, 'xml')
            items = soup.find_all('item')
            
            for i, item in enumerate(items):
                data = {
                    "title": item.find('title').text if item.find('title') else "",
                    "link": item.find('link').text if item.find('link') else "",
                    "description": item.find('description').text if item.find('description') else ""
                }
                
                record = DataRecord(
                    id=f"rss_record_{int(time.time())}_{i}",
                    source_url=url,
                    data=data,
                    extracted_at=time.time(),
                    quality_score=0.6,
                    validation_errors=[]
                )
                records.append(record)
                
        except Exception as e:
            self.logger.error(f"Error en extracción RSS simplificada: {e}")
            raise
        
        return records
    
    def _extract_from_file(self, source_config: Dict[str, Any], **kwargs) -> List[DataRecord]:
        """Extrae datos desde archivos"""
        
        records = []
        
        try:
            file_url = source_config.get("url")
            file_type = source_config.get("file_type", "auto")
            
            # Descargar archivo
            response = requests.get(file_url, timeout=self.config["timeout_seconds"])
            response.raise_for_status()
            
            # Parsear según tipo
            if file_type == "json" or file_url.endswith('.json'):
                data = response.json()
            elif file_type == "csv" or file_url.endswith('.csv'):
                import io
                csv_data = io.StringIO(response.text)
                reader = csv.DictReader(csv_data)
                data = list(reader)
            else:
                # Por defecto, tratar como texto
                data = [{"content": response.text}]
            
            # Convertir a registros
            if isinstance(data, dict):
                data = [data]
            elif not isinstance(data, list):
                data = [{"content": str(data)}]
            
            for i, item in enumerate(data):
                record = DataRecord(
                    id=f"file_record_{int(time.time())}_{i}",
                    source_url=file_url,
                    data=item if isinstance(item, dict) else {"content": str(item)},
                    extracted_at=time.time(),
                    quality_score=0.8,
                    validation_errors=[]
                )
                records.append(record)
                
        except Exception as e:
            self.logger.error(f"Error extrayendo desde archivo: {e}")
            raise
        
        return records
    
    def _validate_source_config(self, config: Dict[str, Any]) -> bool:
        """Valida la configuración de fuente"""
        
        required_fields = ["type", "url"]
        for field in required_fields:
            if field not in config:
                return False
        
        # Validaciones específicas por tipo
        source_type = config.get("type")
        if source_type == "web_api" and "url" not in config:
            return False
        elif source_type == "web_scraping" and "selectors" not in config:
            return False
        
        return True
    
    def _validate_and_clean_records(self, records: List[DataRecord]) -> List[DataRecord]:
        """Valida y limpia registros de datos"""
        
        validated_records = []
        
        for record in records:
            validation_errors = []
            
            # Validaciones básicas
            if not record.data:
                validation_errors.append("Empty data")
            
            # Validar tipos de datos específicos
            for key, value in record.data.items():
                if isinstance(value, str):
                    # Limpiar strings
                    cleaned_value = value.strip()
                    if cleaned_value != value:
                        record.data[key] = cleaned_value
                
                # Validar formato de email si es relevante
                if "email" in key.lower():
                    email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
                    if not email_pattern.match(str(value)):
                        validation_errors.append(f"Invalid email format: {value}")
            
            # Actualizar record con errores de validación
            record.validation_errors = validation_errors
            record.quality_score = self._calculate_quality_score(record)
            
            # Solo incluir registros con calidad mínima
            if record.quality_score >= self.config["quality_threshold"]:
                validated_records.append(record)
        
        return validated_records
    
    def _calculate_quality_score(self, record: DataRecord) -> float:
        """Calcula score de calidad para un registro"""
        
        score = 1.0
        
        # Penalizar por errores de validación
        score -= len(record.validation_errors) * 0.1
        
        # Penalizar por datos vacíos o None
        empty_fields = sum(1 for value in record.data.values() if not value or value == "None")
        total_fields = len(record.data)
        if total_fields > 0:
            completeness_ratio = (total_fields - empty_fields) / total_fields
            score *= completeness_ratio
        
        # Bonificar por consistencia de tipos de datos
        consistent_types = 0
        total_fields = len(record.data)
        for value in record.data.values():
            if value is not None and value != "":
                consistent_types += 1
        
        if total_fields > 0:
            type_consistency = consistent_types / total_fields
            score = (score + type_consistency) / 2
        
        return max(0.0, min(1.0, score))
    
    def _assess_data_quality(self, records: List[DataRecord]) -> DataQuality:
        """Evalúa la calidad general del dataset"""
        
        if not records:
            return DataQuality.INVALID
        
        # Calcular métricas de calidad
        avg_quality = sum(record.quality_score for record in records) / len(records)
        error_rate = sum(len(record.validation_errors) for record in records) / len(records)
        completeness_rate = sum(
            sum(1 for value in record.data.values() if value and value != "None")
            for record in records
        ) / sum(len(record.data) for record in records)
        
        # Determinar nivel de calidad
        if avg_quality >= 0.9 and error_rate <= 0.05 and completeness_rate >= 0.95:
            return DataQuality.EXCELLENT
        elif avg_quality >= 0.8 and error_rate <= 0.1 and completeness_rate >= 0.9:
            return DataQuality.GOOD
        elif avg_quality >= 0.7 and error_rate <= 0.2 and completeness_rate >= 0.8:
            return DataQuality.FAIR
        elif avg_quality >= 0.5:
            return DataQuality.POOR
        else:
            return DataQuality.INVALID
    
    def _infer_schema(self, records: List[DataRecord]) -> Dict[str, str]:
        """Infiere esquema de datos basado en registros"""
        
        if not records:
            return {}
        
        schema = {}
        
        # Recopilar todos los campos únicos
        all_fields = set()
        for record in records:
            all_fields.update(record.data.keys())
        
        # Inferir tipo para cada campo
        for field in all_fields:
            # Analizar valores del campo en todos los registros
            field_values = []
            for record in records:
                if field in record.data:
                    value = record.data[field]
                    if value is not None and value != "":
                        field_values.append(value)
            
            # Inferir tipo basado en valores
            if not field_values:
                schema[field] = "string"  # Default
            elif all(isinstance(v, bool) for v in field_values):
                schema[field] = "boolean"
            elif all(isinstance(v, int) for v in field_values):
                schema[field] = "integer"
            elif all(isinstance(v, float) for v in field_values):
                schema[field] = "float"
            elif all(isinstance(v, str) and re.match(r'\d{4}-\d{2}-\d{2}', str(v)) for v in field_values):
                schema[field] = "date"
            elif all(isinstance(v, str) and re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', str(v)) for v in field_values):
                schema[field] = "email"
            else:
                schema[field] = "string"
        
        return schema
    
    def _analyze_schema(self, schema: Dict[str, str]) -> Dict[str, Any]:
        """Analiza el esquema de datos"""
        
        if not schema:
            return {"error": "No schema available"}
        
        field_types = {}
        for field, field_type in schema.items():
            field_types[field_type] = field_types.get(field_type, 0) + 1
        
        return {
            "total_fields": len(schema),
            "field_types": field_types,
            "most_common_type": max(field_types.items(), key=lambda x: x[1])[0],
            "data_complexity": "high" if len(set(schema.values())) > 3 else "medium"
        }
    
    def _detailed_quality_analysis(self, records: List[DataRecord]) -> Dict[str, Any]:
        """Análisis detallado de calidad de datos"""
        
        if not records:
            return {"error": "No records to analyze"}
        
        # Métricas básicas
        total_records = len(records)
        avg_quality = sum(record.quality_score for record in records) / total_records
        total_errors = sum(len(record.validation_errors) for record in records)
        
        # Análisis de completitud por campo
        field_completeness = {}
        all_fields = set()
        for record in records:
            all_fields.update(record.data.keys())
        
        for field in all_fields:
            non_empty_count = sum(
                1 for record in records
                if field in record.data and record.data[field] and record.data[field] != "None"
            )
            field_completeness[field] = non_empty_count / total_records
        
        # Análisis de tipos de errores
        error_types = {}
        for record in records:
            for error in record.validation_errors:
                error_type = error.split(":")[0] if ":" in error else error
                error_types[error_type] = error_types.get(error_type, 0) + 1
        
        return {
            "total_records": total_records,
            "average_quality_score": avg_quality,
            "total_validation_errors": total_records,
            "error_rate": total_errors / total_records,
            "field_completeness": field_completeness,
            "error_types": error_types,
            "quality_distribution": self._calculate_quality_distribution(records)
        }
    
    def _calculate_quality_distribution(self, records: List[DataRecord]) -> Dict[str, int]:
        """Calcula distribución de scores de calidad"""
        
        distribution = {"excellent": 0, "good": 0, "fair": 0, "poor": 0, "invalid": 0}
        
        for record in records:
            score = record.quality_score
            if score >= 0.9:
                distribution["excellent"] += 1
            elif score >= 0.8:
                distribution["good"] += 1
            elif score >= 0.7:
                distribution["fair"] += 1
            elif score >= 0.5:
                distribution["poor"] += 1
            else:
                distribution["invalid"] += 1
        
        return distribution
    
    def _statistical_analysis(self, records: List[DataRecord]) -> Dict[str, Any]:
        """Análisis estadístico básico"""
        
        if not records:
            return {"error": "No records for statistical analysis"}
        
        # Recopilar datos numéricos
        numeric_fields = {}
        for record in records:
            for field, value in record.data.items():
                if isinstance(value, (int, float)) and value is not None:
                    if field not in numeric_fields:
                        numeric_fields[field] = []
                    numeric_fields[field].append(value)
        
        stats = {}
        for field, values in numeric_fields.items():
            if values:
                stats[field] = {
                    "count": len(values),
                    "mean": sum(values) / len(values),
                    "min": min(values),
                    "max": max(values),
                    "std": self._calculate_std(values)
                }
        
        return {
            "numeric_fields": stats,
            "total_records": len(records),
            "fields_analyzed": len(numeric_fields)
        }
    
    def _calculate_std(self, values: List[float]) -> float:
        """Calcula desviación estándar"""
        if len(values) < 2:
            return 0.0
        
        mean = sum(values) / len(values)
        variance = sum((x - mean) ** 2 for x in values) / len(values)
        return variance ** 0.5
    
    def _detect_patterns(self, records: List[DataRecord]) -> Dict[str, Any]:
        """Detecta patrones en los datos"""
        
        patterns = {
            "temporal_patterns": self._detect_temporal_patterns(records),
            "frequency_patterns": self._detect_frequency_patterns(records),
            "correlation_patterns": self._detect_correlation_patterns(records)
        }
        
        return patterns
    
    def _detect_temporal_patterns(self, records: List[DataRecord]) -> Dict[str, Any]:
        """Detecta patrones temporales"""
        
        timestamps = [record.extracted_at for record in records]
        
        if len(timestamps) < 2:
            return {"pattern": "insufficient_data"}
        
        # Calcular intervalos entre extracciones
        intervals = [timestamps[i] - timestamps[i-1] for i in range(1, len(timestamps))]
        
        return {
            "pattern": "regular" if len(set(int(intervals))) <= 2 else "irregular",
            "avg_interval": sum(intervals) / len(intervals),
            "intervals": intervals[:10]  # Primeros 10 intervalos
        }
    
    def _detect_frequency_patterns(self, records: List[DataRecord]) -> Dict[str, Any]:
        """Detecta patrones de frecuencia"""
        
        # Analizar frecuencia de valores en campos categóricos
        categorical_fields = {}
        
        for record in records:
            for field, value in record.data.items():
                if isinstance(value, str) and field not in categorical_fields:
                    categorical_fields[field] = []
                if field in categorical_fields:
                    categorical_fields[field].append(value)
        
        field_frequencies = {}
        for field, values in categorical_fields.items():
            value_counts = {}
            for value in values:
                value_counts[value] = value_counts.get(value, 0) + 1
            
            # Encontrar el valor más frecuente
            most_common = max(value_counts.items(), key=lambda x: x[1])
            field_frequencies[field] = {
                "most_common_value": most_common[0],
                "frequency": most_common[1] / len(values),
                "unique_values": len(value_counts)
            }
        
        return field_frequencies
    
    def _detect_correlation_patterns(self, records: List[DataRecord]) -> Dict[str, Any]:
        """Detecta patrones de correlación"""
        
        # Análisis simplificado de correlaciones
        correlations = {}
        
        # Identificar campos numéricos
        numeric_pairs = []
        if records:
            first_record = records[0]
            numeric_fields = [k for k, v in first_record.data.items() if isinstance(v, (int, float))]
            
            # Calcular correlaciones simples entre campos numéricos
            for i, field1 in enumerate(numeric_fields):
                for field2 in numeric_fields[i+1:]:
                    values1 = [r.data.get(field1, 0) for r in records if isinstance(r.data.get(field1), (int, float))]
                    values2 = [r.data.get(field2, 0) for r in records if isinstance(r.data.get(field2), (int, float))]
                    
                    if len(values1) == len(values2) and len(values1) > 1:
                        correlation = self._calculate_correlation(values1, values2)
                        correlations[f"{field1}_vs_{field2}"] = correlation
        
        return correlations
    
    def _calculate_correlation(self, x: List[float], y: List[float]) -> float:
        """Calcula correlación de Pearson"""
        if len(x) != len(y) or len(x) < 2:
            return 0.0
        
        n = len(x)
        sum_x = sum(x)
        sum_y = sum(y)
        sum_xy = sum(x[i] * y[i] for i in range(n))
        sum_x2 = sum(xi ** 2 for xi in x)
        sum_y2 = sum(yi ** 2 for yi in y)
        
        numerator = n * sum_xy - sum_x * sum_y
        denominator = ((n * sum_x2 - sum_x ** 2) * (n * sum_y2 - sum_y ** 2)) ** 0.5
        
        return numerator / denominator if denominator != 0 else 0.0
    
    def _analyze_completeness(self, records: List[DataRecord]) -> Dict[str, Any]:
        """Analiza completitud de datos"""
        
        if not records:
            return {"error": "No records to analyze"}
        
        total_records = len(records)
        
        # Calcular completitud por campo
        all_fields = set()
        for record in records:
            all_fields.update(record.data.keys())
        
        field_completeness = {}
        for field in all_fields:
            complete_records = sum(
                1 for record in records
                if field in record.data and record.data[field] and record.data[field] != "None"
            )
            field_completeness[field] = {
                "complete_count": complete_records,
                "completeness_rate": complete_records / total_records,
                "missing_count": total_records - complete_records
            }
        
        # Completitud general
        general_completeness = sum(field["completeness_rate"] for field in field_completeness.values()) / len(field_completeness)
        
        return {
            "overall_completeness": general_completeness,
            "field_completeness": field_completeness,
            "total_fields": len(all_fields),
            "total_records": total_records
        }
    
    def _generate_analysis_recommendations(self, dataset: DataSet) -> List[str]:
        """Genera recomendaciones basadas en análisis"""
        
        recommendations = []
        
        # Recomendaciones basadas en calidad
        if dataset.quality_assessment == DataQuality.POOR:
            recommendations.append("Considerar mejorar la fuente de datos o aplicar filtros de calidad más estrictos")
        
        # Recomendaciones basadas en completitud
        analysis = self.analyze_dataset(dataset)
        completeness = analysis.get("data_completeness", {})
        overall_completeness = completeness.get("overall_completeness", 0)
        
        if overall_completeness < 0.8:
            recommendations.append("La completitud de datos es baja - considerar fuentes adicionales o métodos de imputación")
        
        # Recomendaciones basadas en errores
        quality_analysis = analysis.get("data_quality_analysis", {})
        error_rate = quality_analysis.get("error_rate", 0)
        
        if error_rate > 0.1:
            recommendations.append("Alto índice de errores de validación - revisar y mejorar procesos de extracción")
        
        # Recomendaciones basadas en patrones
        patterns = analysis.get("pattern_detection", {})
        if patterns.get("temporal_patterns", {}).get("pattern") == "irregular":
            recommendations.append("Patrones temporales irregulares detectados - verificar regularidad de la fuente")
        
        return recommendations
    
    def _apply_transformation(self, record: DataRecord, transform: Dict[str, Any]) -> DataRecord:
        """Aplica una transformación específica a un registro"""
        
        transform_type = transform.get("type")
        
        if transform_type == "field_mapping":
            # Mapear campos
            field_mapping = transform.get("field_mapping", {})
            for old_field, new_field in field_mapping.items():
                if old_field in record.data:
                    record.data[new_field] = record.data[old_field]
                    del record.data[old_field]
        
        elif transform_type == "data_type_conversion":
            # Convertir tipos de datos
            conversions = transform.get("conversions", {})
            for field, target_type in conversions.items():
                if field in record.data:
                    value = record.data[field]
                    try:
                        if target_type == "integer":
                            record.data[field] = int(value)
                        elif target_type == "float":
                            record.data[field] = float(value)
                        elif target_type == "boolean":
                            record.data[field] = bool(value)
                        elif target_type == "string":
                            record.data[field] = str(value)
                    except (ValueError, TypeError):
                        pass  # Mantener valor original si la conversión falla
        
        elif transform_type == "data_cleaning":
            # Limpiar datos
            cleaning_rules = transform.get("cleaning_rules", {})
            for field, rules in cleaning_rules.items():
                if field in record.data:
                    value = str(record.data[field])
                    
                    if rules.get("remove_whitespace", False):
                        value = value.strip()
                    if rules.get("lowercase", False):
                        value = value.lower()
                    if rules.get("uppercase", False):
                        value = value.upper()
                    if rules.get("remove_special_chars", False):
                        value = re.sub(r'[^\w\s]', '', value)
                    
                    record.data[field] = value
        
        elif transform_type == "data_validation":
            # Aplicar validaciones
            validations = transform.get("validations", {})
            for field, rules in validations.items():
                if field in record.data:
                    value = record.data[field]
                    
                    # Validar formato de email
                    if rules.get("email_format", False):
                        email_pattern = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
                        if not email_pattern.match(str(value)):
                            record.validation_errors.append(f"Invalid email format: {value}")
                    
                    # Validar rango numérico
                    if isinstance(value, (int, float)):
                        if "min_value" in rules and value < rules["min_value"]:
                            record.validation_errors.append(f"Value {value} below minimum {rules['min_value']}")
                        if "max_value" in rules and value > rules["max_value"]:
                            record.validation_errors.append(f"Value {value} above maximum {rules['max_value']}")
        
        return record
    
    def _export_to_json(self, dataset: DataSet, output_path: str, **kwargs) -> str:
        """Exporta dataset a JSON"""
        
        export_data = {
            "dataset_info": {
                "name": dataset.name,
                "description": dataset.description,
                "source_type": dataset.source_type.value,
                "created_at": dataset.created_at,
                "last_updated": dataset.last_updated
            },
            "schema": dataset.schema,
            "records": [asdict(record) for record in dataset.records]
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
        
        return output_path
    
    def _export_to_csv(self, dataset: DataSet, output_path: str, **kwargs) -> str:
        """Exporta dataset a CSV"""
        
        if not dataset.records:
            raise ValueError("No records to export")
        
        # Recopilar todos los campos únicos
        all_fields = set()
        for record in dataset.records:
            all_fields.update(record.data.keys())
        
        fieldnames = sorted(all_fields)
        
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for record in dataset.records:
                row = {}
                for field in fieldnames:
                    row[field] = record.data.get(field, "")
                writer.writerow(row)
        
        return output_path
    
    def _export_to_xml(self, dataset: DataSet, output_path: str, **kwargs) -> str:
        """Exporta dataset a XML"""
        
        root = ET.Element("dataset")
        root.set("name", dataset.name)
        root.set("source_type", dataset.source_type.value)
        
        # Información del dataset
        info_elem = ET.SubElement(root, "info")
        for key, value in [
            ("description", dataset.description),
            ("created_at", str(dataset.created_at)),
            ("last_updated", str(dataset.last_updated))
        ]:
            elem = ET.SubElement(info_elem, key)
            elem.text = str(value)
        
        # Esquema
        schema_elem = ET.SubElement(root, "schema")
        for field, field_type in dataset.schema.items():
            field_elem = ET.SubElement(schema_elem, "field")
            field_elem.set("name", field)
            field_elem.set("type", field_type)
        
        # Registros
        records_elem = ET.SubElement(root, "records")
        for record in dataset.records:
            record_elem = ET.SubElement(records_elem, "record")
            record_elem.set("id", record.id)
            record_elem.set("source_url", record.source_url)
            record_elem.set("extracted_at", str(record.extracted_at))
            record_elem.set("quality_score", str(record.quality_score))
            
            for key, value in record.data.items():
                data_elem = ET.SubElement(record_elem, key)
                data_elem.text = str(value)
        
        tree = ET.ElementTree(root)
        tree.write(output_path, encoding='utf-8', xml_declaration=True)
        
        return output_path
    
    def _export_to_excel(self, dataset: DataSet, output_path: str, **kwargs) -> str:
        """Exporta dataset a Excel"""
        
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl no está disponible. Instalar con: pip install openpyxl")
        
        # Preparar datos para Excel
        if not dataset.records:
            raise ValueError("No records to export")
        
        # Recopilar todos los campos únicos
        all_fields = set()
        for record in dataset.records:
            all_fields.update(record.data.keys())
        
        fieldnames = sorted(all_fields)
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = dataset.name[:31]  # Limitar nombre de hoja
        
        # Escribir encabezados
        for col, field in enumerate(fieldnames, 1):
            ws.cell(row=1, column=col, value=field)
        
        # Escribir datos
        for row_idx, record in enumerate(dataset.records, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                ws.cell(row=row_idx, column=col_idx, value=record.data.get(field, ""))
        
        # Agregar hoja de información
        info_ws = wb.create_sheet("Dataset_Info")
        info_data = [
            ("Name", dataset.name),
            ("Description", dataset.description),
            ("Source Type", dataset.source_type.value),
            ("Total Records", dataset.total_records),
            ("Created At", datetime.fromtimestamp(dataset.created_at).isoformat()),
            ("Last Updated", datetime.fromtimestamp(dataset.last_updated).isoformat())
        ]
        
        for row_idx, (key, value) in enumerate(info_data, 1):
            info_ws.cell(row=row_idx, column=1, value=key)
            info_ws.cell(row=row_idx, column=2, value=str(value))
        
        wb.save(output_path)
        return output_path
    
    def _export_to_database(self, dataset: DataSet, output_path: str, **kwargs) -> str:
        """Exporta dataset a base de datos SQLite"""
        
        conn = sqlite3.connect(output_path)
        
        try:
            # Preparar datos para inserción
            if not dataset.records:
                raise ValueError("No records to export")
            
            # Recopilar todos los campos únicos
            all_fields = set()
            for record in dataset.records:
                all_fields.update(record.data.keys())
            
            fieldnames = sorted(all_fields)
            
            # Crear tabla
            create_table_sql = f"""
                CREATE TABLE IF NOT EXISTS {dataset.name.replace(' ', '_').lower()} (
                    id TEXT PRIMARY KEY,
                    source_url TEXT,
                    extracted_at REAL,
                    quality_score REAL
                """
            
            for field in fieldnames:
                create_table_sql += f", {field} TEXT"
            
            create_table_sql += ")"
            
            conn.execute(create_table_sql)
            
            # Insertar datos
            for record in dataset.records:
                values = [record.id, record.source_url, record.extracted_at, record.quality_score]
                placeholders = ", ".join(["?" for _ in values])
                field_placeholders = ", ".join(["?" for _ in fieldnames])
                
                insert_sql = f"""
                    INSERT OR REPLACE INTO {dataset.name.replace(' ', '_').lower()} 
                    (id, source_url, extracted_at, quality_score, {", ".join(fieldnames)})
                    VALUES ({placeholders}, {field_placeholders})
                """
                
                field_values = [record.data.get(field, "") for field in fieldnames]
                conn.execute(insert_sql, values + field_values)
            
            conn.commit()
            
        except Exception as e:
            conn.rollback()
            raise
        finally:
            conn.close()
        
        return output_path
    
    def _generate_cache_key(self, source_config: Dict[str, Any]) -> str:
        """Genera clave de cache para configuración de fuente"""
        
        config_str = json.dumps(source_config, sort_keys=True)
        return hashlib.md5(config_str.encode()).hexdigest()
    
    def _init_database(self):
        """Inicializa base de datos interna"""
        
        # Crear directorio de datos si no existe
        import os
        data_dir = os.path.join(os.path.dirname(__file__), '..', '..', '..', 'data')
        os.makedirs(data_dir, exist_ok=True)
        
        # Base de datos para jobs programados y datasets guardados
        self._db_path = os.path.join(data_dir, 'data_mining.db')
        
        # Crear tablas si no existen
        conn = sqlite3.connect(self._db_path)
        try:
            # Tabla de datasets guardados
            conn.execute("""
                CREATE TABLE IF NOT EXISTS saved_datasets (
                    id TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    config_json TEXT NOT NULL,
                    data_json TEXT NOT NULL,
                    created_at REAL NOT NULL,
                    last_accessed REAL NOT NULL
                )
            """)
            
            # Tabla de jobs programados
            conn.execute("""
                CREATE TABLE IF NOT EXISTS scheduled_jobs (
                    job_id TEXT PRIMARY KEY,
                    source_config_json TEXT NOT NULL,
                    schedule_config_json TEXT NOT NULL,
                    status TEXT NOT NULL,
                    created_at REAL NOT NULL,
                    last_run_at REAL,
                    next_run_at REAL
                )
            """)
            
            conn.commit()
        except Exception as e:
            self.logger.error(f"Error inicializando base de datos: {e}")
        finally:
            conn.close()
    
    def get_data_analytics(self, dataset: DataSet) -> Dict[str, Any]:
        """Obtiene analytics detallados del dataset"""
        
        return {
            "dataset_info": {
                "name": dataset.name,
                "source_type": dataset.source_type.value,
                "total_records": dataset.total_records,
                "quality_assessment": dataset.quality_assessment.value,
                "created_at": dataset.created_at,
                "last_updated": dataset.last_updated
            },
            "schema_info": {
                "total_fields": len(dataset.schema),
                "field_types": dict(dataset.schema)
            },
            "quality_metrics": {
                "average_quality_score": sum(r.quality_score for r in dataset.records) / len(dataset.records) if dataset.records else 0,
                "validation_errors": sum(len(r.validation_errors) for r in dataset.records)
            },
            "extraction_metadata": dataset.extraction_config
        }
    
    def clear_extraction_cache(self):
        """Limpia cache de extracciones"""
        self._extraction_cache.clear()
        self._schema_cache.clear()
    
    def get_supported_formats(self) -> List[Dict[str, str]]:
        """Obtiene formatos de datos soportados"""
        return [
            {
                "name": fmt.value,
                "display_name": fmt.name.title(),
                "description": self._get_format_description(fmt)
            }
            for fmt in DataFormat
        ]
    
    def _get_format_description(self, format_type: DataFormat) -> str:
        """Obtiene descripción de un formato"""
        descriptions = {
            DataFormat.JSON: "Formato JSON estructurado con metadatos completos",
            DataFormat.CSV: "Archivo CSV para análisis en hojas de cálculo",
            DataFormat.XML: "Formato XML para interoperabilidad",
            DataFormat.EXCEL: "Archivo Excel con múltiples hojas",
            DataFormat.DATABASE: "Base de datos SQLite para consultas SQL",
            DataFormat.PARQUET: "Formato Parquet optimizado para big data"
        }
        return descriptions.get(format_type, "Formato de datos")


# Funciones de utilidad para compatibilidad MCP
def create_data_mining_agent() -> DataMiningAgent:
    """Crea una instancia del agente de data mining"""
    return DataMiningAgent()


# Testing y demostración
if __name__ == "__main__":
    # Configurar logging
    logging.basicConfig(level=logging.INFO)
    
    # Crear agente
    agent = DataMiningAgent()
    
    print("⛏️ Data Mining Agent - Agente de Extracción de Datos Avanzada")
    print("=" * 65)
    
    # Ejemplo 1: Extracción desde API JSON
    print("\n🔌 Ejemplo 1: Extracción desde API JSON")
    
    api_config = {
        "name": "Posts API Example",
        "description": "Ejemplo de extracción desde API REST",
        "type": "web_api",
        "url": "https://jsonplaceholder.typicode.com/posts",
        "params": {"_limit": "10"}
    }
    
    try:
        dataset = agent.extract_data(api_config, enable_validation=True)
        print(f"  ✅ Extraídos {dataset.total_records} registros")
        print(f"  📊 Calidad: {dataset.quality_assessment.value}")
        print(f"  🏗️ Campos: {len(dataset.schema)}")
        
        # Análisis del dataset
        analysis = agent.analyze_dataset(dataset)
        print(f"  📈 Completitud general: {analysis['data_completeness']['overall_completeness']:.2f}")
        
    except Exception as e:
        print(f"  ❌ Error: {e}")
    
    # Ejemplo 2: Scraping web simplificado
    print("\n🌐 Ejemplo 2: Extracción desde sitio web")
    
    web_config = {
        "name": "Example Web Scraping",
        "description": "Ejemplo de scraping web",
        "type": "web_scraping",
        "url": "https://httpbin.org/html",
        "selectors": {
            "title": {
                "css_selector": "h1",
                "attributes": {}
            },
            "paragraphs": {
                "css_selector": "p"
            }
        }
    }
    
    try:
        dataset = agent.extract_data(web_config, enable_validation=True)
        print(f"  ✅ Extraídos {dataset.total_records} registros")
        print(f"  📊 Calidad: {dataset.quality_assessment.value}")
        
    except Exception as e:
        print(f"  ❌ Error: {e}")
    
    # Ejemplo 3: Análisis detallado
    print("\n📊 Ejemplo 3: Análisis detallado de dataset")
    
    try:
        if 'dataset' in locals() and dataset:
            analysis = agent.analyze_dataset(dataset)
            print(f"  📋 Información del dataset:")
            print(f"    - Nombre: {analysis['dataset_info']['name']}")
            print(f"    - Registros: {analysis['dataset_info']['total_records']}")
            print(f"    - Campos: {analysis['schema_analysis']['total_fields']}")
            print(f"    - Complejidad: {analysis['schema_analysis']['data_complexity']}")
            
            # Recomendaciones
            recommendations = analysis.get('recommendations', [])
            if recommendations:
                print(f"  💡 Recomendaciones:")
                for rec in recommendations[:3]:
                    print(f"    - {rec}")
        
    except Exception as e:
        print(f"  ❌ Error en análisis: {e}")
    
    # Ejemplo 4: Exportación
    print("\n💾 Ejemplo 4: Exportación de datos")
    
    try:
        if 'dataset' in locals() and dataset:
            import tempfile
            
            # Exportar a JSON
            with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
                json_path = agent.export_dataset(dataset, f.name, DataFormat.JSON)
                print(f"  ✅ JSON exportado: {json_path}")
            
            # Exportar a CSV
            with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
                csv_path = agent.export_dataset(dataset, f.name, DataFormat.CSV)
                print(f"  ✅ CSV exportado: {csv_path}")
            
            # Mostrar formatos soportados
            print(f"\n🔧 Formatos de exportación soportados:")
            for fmt in agent.get_supported_formats():
                print(f"  - {fmt['display_name']}: {fmt['description']}")
        
    except Exception as e:
        print(f"  ❌ Error en exportación: {e}")
    
    print(f"\n🎯 Capacidades del Data Mining Agent:")
    print(f"  ✅ Extracción multi-fuente (APIs, web scraping, RSS, archivos)")
    print(f"  ✅ Validación y limpieza automática de datos")
    print(f"  ✅ Análisis estadístico y detección de patrones")
    print(f"  ✅ Exportación en múltiples formatos")
    print(f"  ✅ Evaluación de calidad de datos")
    print(f"  ✅ Programación de extracciones periódicas")