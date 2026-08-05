"""
Document Creation Agent MCP - Agente de Creación de Documentos
Integra con servicios de documentos para crear, formatear y exportar documentos,
hojas de cálculo y presentaciones empresariales.

Autor: Document Creation Agent
Versión: 1.0.0
"""

import asyncio
import json
import time
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional, Union, Tuple
from dataclasses import dataclass, field
from enum import Enum
import csv
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path

# Importar la estructura base del agente MCP
try:
    from .base_agent_wrapper import BaseAgentWrapper, AgentCapability
except ImportError:
    BaseAgentWrapper = object
    AgentCapability = None


class DocumentType(Enum):
    """Tipos de documentos soportados"""
    WORD = "word"
    PDF = "pdf"
    HTML = "html"
    MARKDOWN = "markdown"
    SPREADSHEET = "spreadsheet"
    PRESENTATION = "presentation"


class DocumentFormat(Enum):
    """Formatos de documento disponibles"""
    DOCX = "docx"
    PDF = "pdf"
    TXT = "txt"
    HTML = "html"
    MD = "md"
    CSV = "csv"
    XLSX = "xlsx"
    PPTX = "pptx"


@dataclass
class DocumentTemplate:
    """Estructura de datos para plantillas de documento"""
    id: str
    name: str
    type: DocumentType
    content: str
    variables: List[str] = field(default_factory=list)
    style_config: Dict[str, Any] = field(default_factory=dict)
    created_at: datetime = field(default_factory=datetime.now)


@dataclass
class DocumentContent:
    """Estructura de datos para contenido de documento"""
    title: str
    content: str
    author: str = ""
    created_at: datetime = field(default_factory=datetime.now)
    modified_at: Optional[datetime] = None
    metadata: Dict[str, Any] = field(default_factory=dict)
    sections: List[Dict[str, str]] = field(default_factory=list)
    images: List[str] = field(default_factory=list)
    tables: List[List[Any]] = field(default_factory=list)


@dataclass
class SpreadsheetData:
    """Estructura de datos para hojas de cálculo"""
    name: str
    sheets: Dict[str, List[List[Any]]] = field(default_factory=dict)
    headers: List[str] = field(default_factory=list)
    data_types: Dict[str, str] = field(default_factory=dict)
    formatting: Dict[str, Any] = field(default_factory=dict)


@dataclass
class DocumentResponse:
    """Respuesta consolidada de creación de documento"""
    success: bool
    document_id: str
    action: str
    timestamp: float
    execution_time: float
    file_path: Optional[str] = None
    details: Dict[str, Any] = field(default_factory=dict)
    error: Optional[str] = None


class DocumentCreationAgent(BaseAgentWrapper if BaseAgentWrapper else object):
    """
    Agente de Creación de Documentos que maneja la creación, 
    formato y exportación de documentos empresariales.
    """
    
    def __init__(self):
        if BaseAgentWrapper:
            super().__init__(
                agent_name="DocumentCreationAgent",
                capabilities=[
                    AgentCapability.DOCUMENT_CREATION if AgentCapability else "document_creation",
                    AgentCapability.SPREADSHEET_CREATION if AgentCapability else "spreadsheet_creation",
                    AgentCapability.DOCUMENT_FORMATTING if AgentCapability else "document_formatting",
                    AgentCapability.TEMPLATE_ENGINE if AgentCapability else "template_engine",
                    AgentCapability.DOCUMENT_EXPORT if AgentCapability else "document_export",
                ],
                max_concurrent=5,
                timeout_seconds=60,
                retry_attempts=2
            )
        
        self.logger = logging.getLogger(__name__)
        self._templates: Dict[str, DocumentTemplate] = {}
        self._documents: Dict[str, DocumentContent] = {}
        self._spreadsheets: Dict[str, SpreadsheetData] = {}
        
        # Directorio de salida
        self.output_dir = Path("/tmp/documents")
        self.output_dir.mkdir(exist_ok=True)
        
        # Cargar plantillas predefinidas
        self._load_predefined_templates()
    
    async def _initialize(self):
        """Inicialización específica del agente"""
        await asyncio.sleep(0.1)
        self.logger.info("Document Creation Agent inicializado")
    
    def _load_predefined_templates(self):
        """Cargar plantillas predefinidas"""
        # Plantilla de reporte
        report_template = DocumentTemplate(
            id="template_report",
            name="Reporte Empresarial",
            type=DocumentType.WORD,
            content="""
# {{report_title}}

**Fecha:** {{report_date}}
**Autor:** {{author}}
**Departamento:** {{department}}

## Resumen Ejecutivo

{{executive_summary}}

## Análisis Detallado

{{detailed_analysis}}

## Conclusiones

{{conclusions}}

## Recomendaciones

{{recommendations}}

---
*Documento generado automáticamente el {{generated_date}}*
            """,
            variables=["report_title", "report_date", "author", "department", 
                      "executive_summary", "detailed_analysis", "conclusions", "recommendations", "generated_date"],
            style_config={
                "font_size": 12,
                "font_family": "Arial",
                "margins": {"top": 2.54, "bottom": 2.54, "left": 3.17, "right": 3.17},
                "headers": {"bold": True, "font_size": 16}
            }
        )
        
        # Plantilla de hoja de datos
        data_template = DocumentTemplate(
            id="template_data_sheet",
            name="Hoja de Datos",
            type=DocumentType.SPREADSHEET,
            content="",
            variables=["sheet_name", "headers", "data_rows"],
            style_config={
                "header_row": {"bold": True, "fill_color": "CCCCCC"},
                "alternating_rows": {"fill_color": "F2F2F2"},
                "auto_width": True
            }
        )
        
        # Plantilla de propuesta
        proposal_template = DocumentTemplate(
            id="template_proposal",
            name="Propuesta Comercial",
            type=DocumentType.WORD,
            content="""
# PROPUESTA COMERCIAL

## Información del Cliente
**Cliente:** {{client_name}}
**Empresa:** {{client_company}}
**Fecha:** {{proposal_date}}
**Vendedor:** {{sales_rep}}

## Descripción del Proyecto

{{project_description}}

## Alcance del Trabajo

{{work_scope}}

## Cronograma

{{project_timeline}}

## Inversión

{{investment_details}}

## Términos y Condiciones

{{terms_conditions}}

---
**Contacto:** {{contact_info}}
            """,
            variables=["client_name", "client_company", "proposal_date", "sales_rep",
                      "project_description", "work_scope", "project_timeline", 
                      "investment_details", "terms_conditions", "contact_info"],
            style_config={
                "logo_placeholder": True,
                "color_scheme": "blue",
                "font_family": "Calibri"
            }
        )
        
        self._templates[report_template.id] = report_template
        self._templates[data_template.id] = data_template
        self._templates[proposal_template.id] = proposal_template
    
    def _substitute_variables(self, text: str, variables: Dict[str, str]) -> str:
        """Sustituir variables en texto"""
        result = text
        for var, value in variables.items():
            result = result.replace(f"{{{{var}}}}", str(value))
            result = result.replace(f"{{var}}", str(value))
        return result
    
    async def create_document(
        self,
        template_id: Optional[str] = None,
        title: str = "",
        content: str = "",
        variables: Optional[Dict[str, str]] = None,
        format_type: DocumentFormat = DocumentFormat.MD
    ) -> DocumentResponse:
        """Crear nuevo documento"""
        start_time = time.time()
        
        try:
            document_id = f"doc_{int(time.time() * 1000)}"
            
            # Si se usa plantilla, aplicar variables
            if template_id and template_id in self._templates:
                template = self._templates[template_id]
                if variables:
                    content = self._substitute_variables(template.content, variables)
                else:
                    content = template.content
            
            # Crear contenido del documento
            document = DocumentContent(
                title=title or f"Documento {document_id}",
                content=content,
                author="Sistema",
                metadata={"format": format_type.value, "template": template_id}
            )
            
            # Generar archivo
            file_path = await self._generate_file(document, format_type)
            
            # Guardar en memoria
            self._documents[document_id] = document
            
            self.logger.info(f"Documento creado: {document_id}")
            
            return DocumentResponse(
                success=True,
                document_id=document_id,
                action="create_document",
                timestamp=time.time(),
                execution_time=time.time() - start_time,
                file_path=str(file_path),
                details={
                    "title": document.title,
                    "format": format_type.value,
                    "content_length": len(content),
                    "template_used": template_id
                }
            )
            
        except Exception as e:
            self.logger.error(f"Error creando documento: {str(e)}")
            return DocumentResponse(
                success=False,
                document_id="",
                action="create_document",
                timestamp=time.time(),
                execution_time=time.time() - start_time,
                error=str(e)
            )
    
    async def create_spreadsheet(
        self,
        name: str,
        headers: List[str],
        data: List[List[Any]],
        sheet_formats: Optional[Dict[str, Any]] = None
    ) -> DocumentResponse:
        """Crear hoja de cálculo"""
        start_time = time.time()
        
        try:
            spreadsheet_id = f"sheet_{int(time.time() * 1000)}"
            
            # Crear datos de hoja de cálculo
            spreadsheet = SpreadsheetData(
                name=name,
                sheets={"Sheet1": data},
                headers=headers,
                formatting=sheet_formats or {}
            )
            
            # Generar archivo
            file_path = await self._generate_spreadsheet_file(spreadsheet)
            
            self._spreadsheets[spreadsheet_id] = spreadsheet
            
            self.logger.info(f"Hoja de cálculo creada: {spreadsheet_id}")
            
            return DocumentResponse(
                success=True,
                document_id=spreadsheet_id,
                action="create_spreadsheet",
                timestamp=time.time(),
                execution_time=time.time() - start_time,
                file_path=str(file_path),
                details={
                    "name": name,
                    "headers": headers,
                    "rows_count": len(data),
                    "columns_count": len(headers)
                }
            )
            
        except Exception as e:
            self.logger.error(f"Error creando hoja de cálculo: {str(e)}")
            return DocumentResponse(
                success=False,
                document_id="",
                action="create_spreadsheet",
                timestamp=time.time(),
                execution_time=time.time() - start_time,
                error=str(e)
            )
    
    async def format_document(
        self,
        document_id: str,
        formatting_options: Dict[str, Any]
    ) -> DocumentResponse:
        """Formatear documento existente"""
        start_time = time.time()
        
        try:
            if document_id not in self._documents:
                raise ValueError(f"Documento no encontrado: {document_id}")
            
            document = self._documents[document_id]
            
            # Aplicar formato (simulado)
            # En implementación real: aplicar estilos, fuentes, etc.
            document.metadata["formatting"] = formatting_options
            document.modified_at = datetime.now()
            
            # Regenerar archivo con formato
            format_type = DocumentFormat(document.metadata.get("format", "md"))
            file_path = await self._generate_file(document, format_type)
            
            self.logger.info(f"Documento formateado: {document_id}")
            
            return DocumentResponse(
                success=True,
                document_id=document_id,
                action="format_document",
                timestamp=time.time(),
                execution_time=time.time() - start_time,
                file_path=str(file_path),
                details={
                    "formatting_applied": formatting_options,
                    "modified_at": document.modified_at.isoformat()
                }
            )
            
        except Exception as e:
            self.logger.error(f"Error formateando documento: {str(e)}")
            return DocumentResponse(
                success=False,
                document_id=document_id,
                action="format_document",
                timestamp=time.time(),
                execution_time=time.time() - start_time,
                error=str(e)
            )
    
    async def export_document(
        self,
        document_id: str,
        target_format: DocumentFormat
    ) -> DocumentResponse:
        """Exportar documento a otro formato"""
        start_time = time.time()
        
        try:
            if document_id not in self._documents:
                raise ValueError(f"Documento no encontrado: {document_id}")
            
            document = self._documents[document_id]
            
            # Generar archivo en formato destino
            file_path = await self._generate_file(document, target_format)
            
            self.logger.info(f"Documento exportado: {document_id} a {target_format.value}")
            
            return DocumentResponse(
                success=True,
                document_id=document_id,
                action="export_document",
                timestamp=time.time(),
                execution_time=time.time() - start_time,
                file_path=str(file_path),
                details={
                    "original_format": document.metadata.get("format"),
                    "target_format": target_format.value,
                    "exported_at": datetime.now().isoformat()
                }
            )
            
        except Exception as e:
            self.logger.error(f"Error exportando documento: {str(e)}")
            return DocumentResponse(
                success=False,
                document_id=document_id,
                action="export_document",
                timestamp=time.time(),
                execution_time=time.time() - start_time,
                error=str(e)
            )
    
    async def _generate_file(self, document: DocumentContent, format_type: DocumentFormat) -> Path:
        """Generar archivo según el formato"""
        file_name = f"{document.title.replace(' ', '_').lower()}_{document.metadata.get('id', 'temp')}"
        file_path = self.output_dir / f"{file_name}.{format_type.value}"
        
        if format_type == DocumentFormat.TXT:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(f"{document.title}\n\n{'-' * len(document.title)}\n\n{document.content}")
        
        elif format_type == DocumentFormat.HTML:
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>{document.title}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 40px; }}
                    h1 {{ color: #333; border-bottom: 2px solid #333; }}
                    h2 {{ color: #666; }}
                    .metadata {{ font-size: 12px; color: #888; }}
                </style>
            </head>
            <body>
                <h1>{document.title}</h1>
                <div class="metadata">
                    <p>Autor: {document.author}</p>
                    <p>Creado: {document.created_at.strftime('%d/%m/%Y %H:%M')}</p>
                </div>
                <div>{document.content.replace(chr(10), '<br>')}</div>
            </body>
            </html>
            """
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
        
        else:  # MD (Markdown) por defecto
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(f"# {document.title}\n\n")
                f.write(f"**Autor:** {document.author}\n")
                f.write(f"**Fecha:** {document.created_at.strftime('%d/%m/%Y %H:%M')}\n\n")
                f.write(f"{document.content}")
        
        return file_path
    
    async def _generate_spreadsheet_file(self, spreadsheet: SpreadsheetData) -> Path:
        """Generar archivo de hoja de cálculo"""
        file_path = self.output_dir / f"{spreadsheet.name.replace(' ', '_').lower()}.xlsx"
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        for sheet_name, data in spreadsheet.sheets.items():
            ws = wb.active if sheet_name == "Sheet1" else wb.create_sheet(sheet_name)
            
            # Agregar headers
            if spreadsheet.headers:
                for col, header in enumerate(spreadsheet.headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Agregar datos
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Ajustar ancho de columnas
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
        return file_path
    
    async def list_templates(self) -> DocumentResponse:
        """Listar plantillas disponibles"""
        start_time = time.time()
        
        try:
            templates_list = []
            for template in self._templates.values():
                templates_list.append({
                    "id": template.id,
                    "name": template.name,
                    "type": template.type.value,
                    "variables_count": len(template.variables),
                    "variables": template.variables
                })
            
            return DocumentResponse(
                success=True,
                document_id="templates_list",
                action="list_templates",
                timestamp=time.time(),
                execution_time=time.time() - start_time,
                details={"templates": templates_list}
            )
            
        except Exception as e:
            self.logger.error(f"Error listando plantillas: {str(e)}")
            return DocumentResponse(
                success=False,
                document_id="",
                action="list_templates",
                timestamp=time.time(),
                execution_time=time.time() - start_time,
                error=str(e)
            )
    
    async def process_request(
        self,
        request: Dict[str, Any],
        context: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Procesar request de creación de documentos
        
        Formatos soportados:
        - create_document: {"action": "create_document", "title": "Mi Documento", "content": "Contenido..."}
        - create_spreadsheet: {"action": "create_spreadsheet", "name": "Ventas", "headers": ["Fecha", "Cantidad"], "data": [["2023-01-01", 100]]}
        - format_document: {"action": "format_document", "document_id": "doc_123", "formatting_options": {...}}
        - export_document: {"action": "export_document", "document_id": "doc_123", "target_format": "pdf"}
        - list_templates: {"action": "list_templates"}
        """
        try:
            await self.ensure_initialized()
            
            action = request.get("action", "").lower()
            
            if action == "create_document":
                template_id = request.get("template_id")
                title = request.get("title", "")
                content = request.get("content", "")
                variables = request.get("variables", {})
                format_str = request.get("format", "md")
                
                try:
                    format_type = DocumentFormat(format_str.lower())
                except ValueError:
                    format_type = DocumentFormat.MD
                
                if AgentCapability:
                    response = await self.execute_operation(
                        operation_name="create_document",
                        capability=AgentCapability.DOCUMENT_CREATION,
                        operation_func=self.create_document,
                        template_id=template_id,
                        title=title,
                        content=content,
                        variables=variables,
                        format_type=format_type
                    )
                else:
                    response = await self.create_document(template_id, title, content, variables, format_type)
                
                return {
                    "success": response.success,
                    "document_id": response.document_id,
                    "file_path": response.file_path,
                    "details": response.details,
                    "error": response.error
                }
            
            elif action == "create_spreadsheet":
                name = request.get("name", "Hoja de Cálculo")
                headers = request.get("headers", [])
                data = request.get("data", [])
                sheet_formats = request.get("sheet_formats", {})
                
                if not headers:
                    raise ValueError("Headers son requeridos para crear hoja de cálculo")
                
                if AgentCapability:
                    response = await self.execute_operation(
                        operation_name="create_spreadsheet",
                        capability=AgentCapability.SPREADSHEET_CREATION,
                        operation_func=self.create_spreadsheet,
                        name=name,
                        headers=headers,
                        data=data,
                        sheet_formats=sheet_formats
                    )
                else:
                    response = await self.create_spreadsheet(name, headers, data, sheet_formats)
                
                return {
                    "success": response.success,
                    "document_id": response.document_id,
                    "file_path": response.file_path,
                    "details": response.details,
                    "error": response.error
                }
            
            elif action == "format_document":
                document_id = request.get("document_id")
                formatting_options = request.get("formatting_options", {})
                
                if not document_id:
                    raise ValueError("document_id requerido")
                
                if AgentCapability:
                    response = await self.execute_operation(
                        operation_name="format_document",
                        capability=AgentCapability.DOCUMENT_FORMATTING,
                        operation_func=self.format_document,
                        document_id=document_id,
                        formatting_options=formatting_options
                    )
                else:
                    response = await self.format_document(document_id, formatting_options)
                
                return {
                    "success": response.success,
                    "document_id": response.document_id,
                    "file_path": response.file_path,
                    "details": response.details,
                    "error": response.error
                }
            
            elif action == "export_document":
                document_id = request.get("document_id")
                target_format = request.get("target_format", "pdf")
                
                if not document_id:
                    raise ValueError("document_id requerido")
                
                try:
                    format_type = DocumentFormat(target_format.lower())
                except ValueError:
                    return {
                        "success": False,
                        "error": f"Formato no soportado: {target_format}"
                    }
                
                if AgentCapability:
                    response = await self.execute_operation(
                        operation_name="export_document",
                        capability=AgentCapability.DOCUMENT_EXPORT,
                        operation_func=self.export_document,
                        document_id=document_id,
                        target_format=format_type
                    )
                else:
                    response = await self.export_document(document_id, format_type)
                
                return {
                    "success": response.success,
                    "document_id": response.document_id,
                    "file_path": response.file_path,
                    "details": response.details,
                    "error": response.error
                }
            
            elif action == "list_templates":
                if AgentCapability:
                    response = await self.execute_operation(
                        operation_name="list_templates",
                        capability=AgentCapability.TEMPLATE_ENGINE,
                        operation_func=self.list_templates
                    )
                else:
                    response = await self.list_templates()
                
                return {
                    "success": response.success,
                    "templates": response.details.get("templates", []) if response.success else [],
                    "error": response.error
                }
            
            else:
                raise ValueError(f"Acción no soportada: {action}")
                
        except Exception as e:
            self.logger.error(f"Error procesando request de documentos: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }
    
    def get_stats(self) -> Dict[str, Any]:
        """Obtener estadísticas del agente"""
        return {
            "total_documents": len(self._documents),
            "total_spreadsheets": len(self._spreadsheets),
            "total_templates": len(self._templates),
            "output_directory": str(self.output_dir),
            "agent_name": "DocumentCreationAgent",
            "supported_formats": [fmt.value for fmt in DocumentFormat],
            "available_actions": [
                "create_document",
                "create_spreadsheet", 
                "format_document",
                "export_document",
                "list_templates"
            ]
        }